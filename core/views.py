"""
View functions for the attendance management system
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse, HttpResponseForbidden
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone
from datetime import datetime
import csv
import logging

from .models import User, Company, AttendanceRecord, UploadLog, RemarkReason, AttendanceRemark
from .decorators import role_required, company_access_required, check_record_company_access, can_edit_record, can_delete_record
from .forms import LoginForm

# Get logger
logger = logging.getLogger('core')


def csrf_failure(request, reason=""):
    """Custom CSRF failure view - logs out user and shows error"""
    from django.contrib.auth import logout
    
    # Log out the user to clear invalid session
    if request.user.is_authenticated:
        username = request.user.username
        logout(request)
        logger.warning(f'User {username} logged out due to CSRF failure: {reason}')
    
    return render(request, '403_csrf.html', status=403)


def login_view(request):
    """Handle user login"""
    if request.user.is_authenticated:
        return redirect('core:dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            logger.info(f'Successful login: {user.username} (Role: {user.role})')
            messages.success(request, f'Welcome back, {user.username}!')
            return redirect('core:dashboard')
        else:
            logger.warning(f'Failed login attempt for username: {username}')
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'login.html')


def logout_view(request):
    """Handle user logout"""
    username = request.user.username if request.user.is_authenticated else 'Unknown'
    logout(request)
    logger.info(f'User logged out: {username}')
    messages.success(request, 'You have been logged out successfully.')
    return redirect('core:login')



@role_required(['root', 'admin'])
def upload_csv_view(request):
    """Handle CSV/Excel file upload and processing"""
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        
        # Validate file extension
        allowed_extensions = ['.csv', '.xls', '.xlsx']
        if not any(csv_file.name.lower().endswith(ext) for ext in allowed_extensions):
            logger.warning(f'Invalid file upload attempt by {request.user.username}: {csv_file.name}')
            messages.error(request, 'Please upload a valid CSV, XLS, or XLSX file.')
            return redirect('core:upload')
        
        logger.info(f'CSV upload started by {request.user.username}: {csv_file.name}')
        
        # Create a unique upload ID for this session
        import json
        import os
        upload_id = f"upload_{request.user.id}_{int(timezone.now().timestamp())}"
        progress_file = os.path.join(os.path.dirname(__file__), '..', 'logs', f'{upload_id}.json')
        
        # Ensure logs directory exists
        os.makedirs(os.path.dirname(progress_file), exist_ok=True)
        
        # Initialize progress in file
        with open(progress_file, 'w') as f:
            json.dump({'processed': 0, 'total': 0, 'status': 'starting', 'percentage': 0}, f)
        
        # Store upload_id in session and save immediately
        request.session['current_upload_id'] = upload_id
        request.session['progress_file'] = progress_file
        request.session.save()
        
        # Process CSV with progress callback
        from .csv_processor import CSVProcessor
        processor = CSVProcessor()
        
        # Set progress callback that updates file in real-time
        def progress_callback(processed, total, current_ep=None):
            try:
                with open(progress_file, 'w') as f:
                    progress_data = {
                        'processed': processed,
                        'total': total,
                        'status': 'processing',
                        'percentage': int((processed / total) * 100) if total > 0 else 0
                    }
                    if current_ep:
                        progress_data['current_ep'] = current_ep
                    json.dump(progress_data, f)
            except Exception as e:
                logger.error(f'Error writing progress: {e}')
        
        processor.progress_callback = progress_callback
        result = processor.process_csv(csv_file, request.user)
        
        # Mark as complete
        try:
            with open(progress_file, 'w') as f:
                json.dump({
                    'processed': result.get('processed_rows', 0),
                    'total': result.get('total_rows', 0),
                    'status': 'complete',
                    'percentage': 100
                }, f)
        except Exception as e:
            logger.error(f'Error writing final progress: {e}')
        
        # Create upload log
        error_messages = '\n'.join(result['errors']) if result['errors'] else ''
        upload_log = UploadLog.objects.create(
            user=request.user,
            filename=csv_file.name,
            success_count=result['success_count'],
            updated_count=result['updated_count'],
            error_count=result['error_count'],
            error_messages=error_messages
        )
        
        # Log results
        logger.info(
            f'CSV processing completed by {request.user.username}: '
            f'Created={result["success_count"]}, Updated={result["updated_count"]}, Errors={result["error_count"]}'
        )
        
        if result['error_count'] > 0:
            logger.error(f'CSV processing errors for {csv_file.name}: {error_messages[:200]}...')
        
        # Display results
        if result['success']:
            messages.success(
                request,
                f'CSV processed successfully! Created: {result["success_count"]}, Updated: {result["updated_count"]}'
            )
        else:
            messages.warning(
                request,
                f'CSV processed with errors. Created: {result["success_count"]}, Updated: {result["updated_count"]}, Errors: {result["error_count"]}'
            )
            if result['errors']:
                for error in result['errors'][:5]:  # Show first 5 errors
                    messages.error(request, error)
                if len(result['errors']) > 5:
                    messages.info(request, f'... and {len(result["errors"]) - 5} more errors. Check upload logs for details.')
        
        return redirect('core:upload')
    
    # Get recent upload logs for current user
    if request.user.role == 'root':
        recent_logs = UploadLog.objects.all()[:10]
    else:
        recent_logs = UploadLog.objects.filter(user__company=request.user.company)[:10]
    
    context = {
        'recent_logs': recent_logs
    }
    return render(request, 'upload.html', context)


@login_required
def dashboard_view(request):
    """Main dashboard view"""
    from .services.access_control_service import AccessControlService
    import calendar
    from datetime import datetime, date
    
    # Get calendar month (this now controls both calendar and all summaries)
    calendar_month = request.GET.get('calendar_month', '').strip()
    
    # Auto-select current month if not provided (default behavior)
    if not calendar_month:
        today = date.today()
        calendar_month = f"{today.year}-{today.month:02d}"
    
    # Parse month for filtering summaries
    calendar_month_name = ''
    calendar_year = ''
    dashboard_year_int = None
    dashboard_month_int = None
    
    try:
        dashboard_year_int, dashboard_month_int = map(int, calendar_month.split('-'))
        calendar_month_name = calendar.month_name[dashboard_month_int]
        calendar_year = dashboard_year_int
    except (ValueError, IndexError):
        today = date.today()
        calendar_month = f"{today.year}-{today.month:02d}"
        dashboard_year_int = today.year
        dashboard_month_int = today.month
        calendar_month_name = calendar.month_name[dashboard_month_int]
        calendar_year = dashboard_year_int
    
    # Calculate statistics based on user role
    if request.user.role == 'root':
        total_records = AttendanceRecord.objects.count()
        total_companies = Company.objects.count()
        recent_uploads = UploadLog.objects.all()[:5]
    elif request.user.role == 'user1':
        # User1: Calculate statistics only for assigned employees
        queryset = AttendanceRecord.objects.filter(company=request.user.company)
        queryset = AccessControlService.filter_queryset_by_access(queryset, request.user)
        total_records = queryset.count()
        total_companies = 1 if request.user.company else 0
        recent_uploads = UploadLog.objects.filter(user__company=request.user.company)[:5]
    else:
        total_records = AttendanceRecord.objects.filter(company=request.user.company).count()
        total_companies = 1 if request.user.company else 0
        recent_uploads = UploadLog.objects.filter(user__company=request.user.company)[:5]
    
    # Handle calendar view
    calendar_data = None
    calendar_ep = request.GET.get('calendar_ep', '').strip()
    calendar_stats = None
    employee_name = None
    
    if calendar_ep:
        try:
            # Use the already parsed month values
            year = dashboard_year_int
            month = dashboard_month_int
            
            # Get attendance records for the EP and month
            if request.user.role == 'root':
                records = AttendanceRecord.objects.filter(
                    ep_no__icontains=calendar_ep,
                    date__year=year,
                    date__month=month
                )
            elif request.user.role == 'user1':
                records = AttendanceRecord.objects.filter(
                    company=request.user.company,
                    ep_no__icontains=calendar_ep,
                    date__year=year,
                    date__month=month
                )
                records = AccessControlService.filter_queryset_by_access(records, request.user)
            else:
                records = AttendanceRecord.objects.filter(
                    company=request.user.company,
                    ep_no__icontains=calendar_ep,
                    date__year=year,
                    date__month=month
                )
            
            # Only generate calendar if records exist
            if records.exists():
                # Calculate statistics
                present_days = records.filter(status='P').count()
                absent_days = records.filter(status='A').count()
                late_days = records.filter(status='L').count()
                partial_days = records.filter(status='PD').count()
                total_logged = records.count()
                total_days_in_month = calendar.monthrange(year, month)[1]
                
                # Calculate exceptions (late + partial + absent)
                exception_days = late_days + partial_days + absent_days
                exception_type = "Days"
                if partial_days > 0:
                    exception_type = "Deduction"
                
                calendar_stats = {
                    'present_days': present_days,
                    'exception_days': exception_days if exception_days > 0 else partial_days,
                    'exception_type': exception_type,
                    'total_logged': total_logged,
                    'total_days': total_days_in_month,
                }
                
                # Create a dictionary of dates to records
                records_dict = {record.date.day: record for record in records}
                
                # Get employee name from first record
                employee_name = records.first().ep_name
                
                # Generate calendar data
                cal = calendar.monthcalendar(year, month)
                calendar_data = []
                
                for week in cal:
                    for day in week:
                        if day == 0:
                            # Day from previous/next month
                            calendar_data.append({
                                'day': '',
                                'is_other_month': True,
                                'status': None,
                                'hours': None,
                                'ep_name': employee_name
                            })
                        else:
                            record = records_dict.get(day)
                            calendar_data.append({
                                'day': day,
                                'is_other_month': False,
                                'status': record.status if record else None,
                                'hours': record.hours if record and record.hours else None,
                                'ep_name': employee_name,
                                'deduction': record.deduction if record and hasattr(record, 'deduction') else None
                            })
                
                # Debug logging
                logger.info(f'Calendar data generated: {len(calendar_data)} days, first 5: {calendar_data[:5]}')
            else:
                # No records found - calendar_data remains None
                logger.info(f'No records found for EP: {calendar_ep}, Month: {calendar_month_name} {calendar_year}')
        except (ValueError, AttributeError):
            pass
    
    # Get EIC pending request statistics
    from django.db.models import Count, Q, Case, When, IntegerField
    eic_pending_requests = []
    grand_total_approved = 0
    grand_total_pending = 0
    grand_total_all = 0
    
    # Partial Day statistics by EIC
    partial_day_by_eic = []
    pd_grand_total = 0
    
    # Regularization statistics by EIC
    regularization_by_eic = []
    reg_total_count = 0
    
    try:
        if request.user.role == 'root':
            eic_queryset = AttendanceRecord.objects.exclude(requested_eic_name='')
            pd_queryset = AttendanceRecord.objects.filter(status='PD').exclude(requested_eic_name='')
            # Regularization: records with in_time_2 or out_time_2 (new punch times) and has status
            reg_queryset = AttendanceRecord.objects.filter(
                Q(in_time_2__isnull=False) | Q(out_time_2__isnull=False)
            ).exclude(requested_eic_name='').exclude(Q(ot_request_status='') | Q(ot_request_status__isnull=True))
        elif request.user.role == 'user1':
            eic_queryset = AttendanceRecord.objects.filter(
                company=request.user.company
            ).exclude(requested_eic_name='')
            eic_queryset = AccessControlService.filter_queryset_by_access(eic_queryset, request.user)
            
            pd_queryset = AttendanceRecord.objects.filter(
                company=request.user.company,
                status='PD'
            ).exclude(requested_eic_name='')
            pd_queryset = AccessControlService.filter_queryset_by_access(pd_queryset, request.user)
            
            reg_queryset = AttendanceRecord.objects.filter(
                company=request.user.company
            ).filter(
                Q(in_time_2__isnull=False) | Q(out_time_2__isnull=False)
            ).exclude(requested_eic_name='').exclude(Q(ot_request_status='') | Q(ot_request_status__isnull=True))
            reg_queryset = AccessControlService.filter_queryset_by_access(reg_queryset, request.user)
        else:
            eic_queryset = AttendanceRecord.objects.filter(
                company=request.user.company
            ).exclude(requested_eic_name='')
            
            pd_queryset = AttendanceRecord.objects.filter(
                company=request.user.company,
                status='PD'
            ).exclude(requested_eic_name='')
            
            reg_queryset = AttendanceRecord.objects.filter(
                company=request.user.company
            ).filter(
                Q(in_time_2__isnull=False) | Q(out_time_2__isnull=False)
            ).exclude(requested_eic_name='').exclude(Q(ot_request_status='') | Q(ot_request_status__isnull=True))
        
        # Apply calendar month filter to all summaries (calendar_month is always set to current month by default)
        if dashboard_year_int and dashboard_month_int:
            eic_queryset = eic_queryset.filter(date__year=dashboard_year_int, date__month=dashboard_month_int)
            pd_queryset = pd_queryset.filter(date__year=dashboard_year_int, date__month=dashboard_month_int)
            reg_queryset = reg_queryset.filter(date__year=dashboard_year_int, date__month=dashboard_month_int)
        
        # Group by EIC name and count by status for OT
        eic_pending_requests = eic_queryset.values('requested_eic_name').annotate(
            approved_count=Count('id', filter=Q(ot_request_status='Approved')),
            pending_count=Count('id', filter=Q(ot_request_status='Pending')),
            total_count=Count('id')
        ).order_by('-total_count')[:10]  # Top 10 EICs with most requests
        
        logger.info(f'EIC Overtime Summary: Found {len(list(eic_pending_requests))} EICs')
        
        # Calculate grand totals for OT
        grand_total_approved = eic_queryset.filter(ot_request_status='Approved').count()
        grand_total_pending = eic_queryset.filter(ot_request_status='Pending').count()
        grand_total_all = eic_queryset.count()
        
        # Group by EIC for Partial Day
        partial_day_by_eic = pd_queryset.values('requested_eic_name').annotate(
            total_count=Count('id')
        ).order_by('-total_count')[:10]
        pd_grand_total = pd_queryset.count()
        
        # Group by EIC for Regularization with status breakdown
        regularization_by_eic = reg_queryset.values('requested_eic_name').annotate(
            approved_count=Count('id', filter=Q(ot_request_status='Approved')),
            pending_count=Count('id', filter=Q(ot_request_status='Pending')),
            total_count=Count('id')
        ).order_by('-total_count')[:10]
        reg_total_count = reg_queryset.count()
        reg_total_approved = reg_queryset.filter(ot_request_status='Approved').count()
        reg_total_pending = reg_queryset.filter(ot_request_status='Pending').count()
        
        # ARC Summary - Pivot table by trade and contractor (top 7 contractors by mandays)
        from django.db.models import Sum, FloatField
        from collections import defaultdict
        
        # Get base queryset for ARC
        if request.user.role == 'root':
            arc_queryset = AttendanceRecord.objects.exclude(cont_code='Unknown').exclude(cont_code='').exclude(trade='').exclude(trade__isnull=True)
        elif request.user.role == 'user1':
            arc_queryset = AttendanceRecord.objects.filter(
                company=request.user.company
            ).exclude(cont_code='Unknown').exclude(cont_code='').exclude(trade='').exclude(trade__isnull=True)
            arc_queryset = AccessControlService.filter_queryset_by_access(arc_queryset, request.user)
        else:
            arc_queryset = AttendanceRecord.objects.filter(
                company=request.user.company
            ).exclude(cont_code='Unknown').exclude(cont_code='').exclude(trade='').exclude(trade__isnull=True)
        
        # Apply calendar month filter to ARC Summary (calendar_month is always set to current month by default)
        if dashboard_year_int and dashboard_month_int:
            arc_queryset = arc_queryset.filter(date__year=dashboard_year_int, date__month=dashboard_month_int)
        
        # Get top 7 contractors by total mandays
        top_contractors = arc_queryset.values('cont_code', 'contract').annotate(
            total_mandays=Sum('mandays', output_field=FloatField())
        ).order_by('-total_mandays')[:7]
        
        # Use contractor names instead of codes
        arc_contractors = [c['contract'] or c['cont_code'] for c in top_contractors]
        # Create mapping from code to name for lookup
        code_to_name = {c['cont_code']: (c['contract'] or c['cont_code']) for c in top_contractors}
        arc_contractor_codes = [c['cont_code'] for c in top_contractors]
        
        # Build pivot data: trade -> {contractor name -> mandays}
        pivot_data = defaultdict(lambda: defaultdict(float))
        trade_totals = defaultdict(float)
        contractor_totals = defaultdict(float)
        
        # Query data for top contractors only
        arc_records = arc_queryset.filter(cont_code__in=arc_contractor_codes).values('trade', 'cont_code', 'contract', 'mandays')
        
        for record in arc_records:
            trade = record['trade']
            if not trade:  # Skip if trade is empty or None
                continue
            cont_code = record['cont_code']
            cont_name = record['contract'] or cont_code
            mandays = float(record['mandays'] or 0)
            
            pivot_data[trade][cont_name] += mandays
            trade_totals[trade] += mandays
            contractor_totals[cont_name] += mandays
        
        # Convert to list format for template
        arc_summary_data = []
        for trade in sorted(pivot_data.keys()):
            row = {
                'trade': trade,
                'contractors': {},
                'total': trade_totals[trade]
            }
            for cont_name in arc_contractors:
                row['contractors'][cont_name] = pivot_data[trade].get(cont_name, 0)
            arc_summary_data.append(row)
        
        # Calculate grand totals
        arc_grand_totals = {cont_name: contractor_totals[cont_name] for cont_name in arc_contractors}
        arc_grand_totals['total'] = sum(contractor_totals.values())
        
    except Exception as e:
        logger.error(f'Error fetching EIC statistics: {e}')
        eic_pending_requests = []
        grand_total_approved = 0
        grand_total_pending = 0
        grand_total_all = 0
        partial_day_by_eic = []
        pd_grand_total = 0
        regularization_by_eic = []
        reg_total_count = 0
        reg_total_approved = 0
        reg_total_pending = 0
        arc_summary_data = []
        arc_contractors = []
        arc_grand_totals = {}
    
    context = {
        'total_records': total_records,
        'total_companies': total_companies,
        'recent_uploads': recent_uploads,
        'calendar_data': calendar_data,
        'calendar_ep': calendar_ep,
        'calendar_month': calendar_month,
        'calendar_month_name': calendar_month_name,
        'calendar_year': calendar_year,
        'calendar_stats': calendar_stats,
        'employee_name': employee_name,
        'eic_pending_requests': eic_pending_requests,
        'grand_total_approved': grand_total_approved,
        'grand_total_pending': grand_total_pending,
        'grand_total_all': grand_total_all,
        'partial_day_by_eic': partial_day_by_eic,
        'pd_grand_total': pd_grand_total,
        'regularization_by_eic': regularization_by_eic,
        'reg_total_count': reg_total_count,
        'reg_total_approved': reg_total_approved,
        'reg_total_pending': reg_total_pending,
        'arc_summary_data': arc_summary_data,
        'arc_contractors': arc_contractors,
        'arc_grand_totals': arc_grand_totals,
        'reg_total_approved': reg_total_approved,
        'reg_total_pending': reg_total_pending,
    }
    return render(request, 'dashboard.html', context)


@login_required
@company_access_required
def attendance_list_view(request):
    """List attendance records with filtering and pagination"""
    from .services.access_control_service import AccessControlService
    from django.db.models import Q
    
    # Base queryset based on user role
    if request.user.role == 'root':
        queryset = AttendanceRecord.objects.all()
    elif request.user.role == 'user1':
        # User1: Filter by company and assigned employees
        queryset = AttendanceRecord.objects.filter(company=request.user.company)
        queryset = AccessControlService.filter_queryset_by_access(queryset, request.user)
        
        # Check if user has no assignments
        assigned_employees = AccessControlService.get_assigned_employees(request.user)
        if assigned_employees is not None and len(assigned_employees) == 0:
            messages.info(request, 'You have no employee assignments. Please request access to employees.')
    else:
        # Admin: Filter by company only
        queryset = AttendanceRecord.objects.filter(company=request.user.company)
    
    # Apply filters with default yesterday date
    from datetime import timedelta
    yesterday = (timezone.now() - timedelta(days=1)).date()
    
    date_from = request.GET.get('date_from', '').strip() or str(yesterday)
    date_to = request.GET.get('date_to', '').strip() or str(yesterday)
    company_id = request.GET.get('company', '').strip()
    ep_no = request.GET.get('ep_no', '').strip()
    status = request.GET.get('status', '').strip()
    overstay_filter = request.GET.get('overstay_filter', '').strip()
    show_incomplete = request.GET.get('show_incomplete', '').strip()
    
    # Filter out incomplete records by default (unless explicitly requested)
    if show_incomplete != 'yes':
        # Exclude records that have status 'P' but missing essential time data
        # These are typically records with only status but no actual punch times
        queryset = queryset.exclude(
            Q(status='P') & 
            Q(in_time__isnull=True) & 
            Q(out_time__isnull=True) &
            Q(in_time_2__isnull=True) & 
            Q(out_time_2__isnull=True)
        )
    
    # Debug logging
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"Filters - EP NO: '{ep_no}', Date From: '{date_from}', Date To: '{date_to}', Status: '{status}', Overstay: '{overstay_filter}'")
    
    # Only apply date filters if values are provided
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
        logger.info(f"Applied date_from filter: {date_from}")
    if date_to:
        queryset = queryset.filter(date__lte=date_to)
        logger.info(f"Applied date_to filter: {date_to}")
    if company_id and request.user.role == 'root':
        queryset = queryset.filter(company_id=company_id)
    if ep_no:
        queryset = queryset.filter(ep_no__icontains=ep_no)
        logger.info(f"Applied ep_no filter: {ep_no}")
    if status:
        queryset = queryset.filter(status=status)
    
    # Apply overstay filter
    if overstay_filter == 'has_overstay':
        # Filter for records with overstay (not empty, not "00:00", not "-")
        queryset = queryset.exclude(Q(overstay='') | Q(overstay='00:00') | Q(overstay='-') | Q(overstay__isnull=True))
        logger.info("Applied has_overstay filter")
    elif overstay_filter == 'no_overstay':
        # Filter for records without overstay (empty, "00:00", or "-")
        queryset = queryset.filter(Q(overstay='') | Q(overstay='00:00') | Q(overstay='-') | Q(overstay__isnull=True))
        logger.info("Applied no_overstay filter")
    elif overstay_filter.startswith('range_'):
        # Filter for records with overstay in a specific range (e.g., range_1_2 for 1-2 hours)
        try:
            parts = overstay_filter.split('_')
            min_hours = int(parts[1])
            max_hours = int(parts[2])
            
            # Get all records with overstay
            all_records = queryset.exclude(Q(overstay='') | Q(overstay='00:00') | Q(overstay='-') | Q(overstay__isnull=True))
            
            # Filter by parsing overstay time
            filtered_ids = []
            for record in all_records.iterator(chunk_size=500):
                try:
                    # Parse overstay format (e.g., "02:30", "1:15", "00:45")
                    overstay_str = str(record.overstay).strip()
                    if ':' in overstay_str:
                        parts = overstay_str.split(':')
                        overstay_hours = int(parts[0])
                        overstay_minutes = int(parts[1]) if len(parts) > 1 else 0
                        total_hours = overstay_hours + (overstay_minutes / 60.0)
                        
                        if min_hours <= total_hours < max_hours:
                            filtered_ids.append(record.id)
                except (ValueError, AttributeError):
                    continue
            
            # Handle filtered results
            if filtered_ids:
                matching_records = list(queryset.filter(pk__in=filtered_ids[:999]).select_related('company'))
                
                if len(filtered_ids) > 999:
                    for i in range(999, len(filtered_ids), 999):
                        batch_ids = filtered_ids[i:i+999]
                        matching_records.extend(list(queryset.filter(pk__in=batch_ids).select_related('company')))
                
                matching_records.sort(key=lambda x: (x.date, x.ep_no), reverse=True)
                
                logger.info(f"Applied overstay range {min_hours}-{max_hours} hours filter, found {len(matching_records)} records")
                
                paginator = Paginator(matching_records, 100)
                page_number = request.GET.get('page')
                page_obj = paginator.get_page(page_number)
                
                for record in page_obj:
                    if record.shift and '(' in str(record.shift):
                        record.shift_code = str(record.shift).split('(')[0].strip()
                    else:
                        record.shift_code = str(record.shift) if record.shift else '-'
                
                companies = Company.objects.all() if request.user.role == 'root' else []
                
                context = {
                    'page_obj': page_obj,
                    'companies': companies,
                    'can_edit': can_edit_record(request.user),
                    'can_delete': can_delete_record(request.user),
                }
                return render(request, 'attendance_list.html', context)
            else:
                queryset = queryset.none()
                logger.info(f"Applied overstay range {min_hours}-{max_hours} hours filter, found 0 records")
        except (ValueError, IndexError):
            logger.warning(f"Invalid overstay range filter format: {overstay_filter}")
    elif overstay_filter.startswith('gt_'):
        # Filter for records with overstay greater than X hours
        try:
            hours = int(overstay_filter.split('_')[1])
            # Get all records with overstay - use iterator to avoid loading all at once
            all_records = queryset.exclude(Q(overstay='') | Q(overstay='00:00') | Q(overstay='-') | Q(overstay__isnull=True))
            
            # Filter by parsing overstay time
            filtered_ids = []
            for record in all_records.iterator(chunk_size=500):
                try:
                    # Parse overstay format (e.g., "02:30", "1:15", "00:45")
                    overstay_str = str(record.overstay).strip()
                    if ':' in overstay_str:
                        parts = overstay_str.split(':')
                        overstay_hours = int(parts[0])
                        overstay_minutes = int(parts[1]) if len(parts) > 1 else 0
                        total_hours = overstay_hours + (overstay_minutes / 60.0)
                        
                        if total_hours > hours:
                            filtered_ids.append(record.id)
                except (ValueError, AttributeError):
                    continue
            
            # For large result sets, fetch the actual records and use Python list
            # This avoids SQLite's 999 variable limit with IN clauses
            if filtered_ids:
                # Get the base queryset and fetch matching records
                matching_records = list(queryset.filter(pk__in=filtered_ids[:999]).select_related('company'))
                
                # If we have more than 999 IDs, fetch in batches
                if len(filtered_ids) > 999:
                    for i in range(999, len(filtered_ids), 999):
                        batch_ids = filtered_ids[i:i+999]
                        matching_records.extend(list(queryset.filter(pk__in=batch_ids).select_related('company')))
                
                # Sort the records
                matching_records.sort(key=lambda x: (x.date, x.ep_no), reverse=True)
                
                logger.info(f"Applied overstay > {hours} hours filter, found {len(matching_records)} records")
                
                # Use manual pagination with the list
                paginator = Paginator(matching_records, 100)
                page_number = request.GET.get('page')
                page_obj = paginator.get_page(page_number)
                
                # Add shift_code attribute to each record
                for record in page_obj:
                    if record.shift and '(' in str(record.shift):
                        record.shift_code = str(record.shift).split('(')[0].strip()
                    else:
                        record.shift_code = str(record.shift) if record.shift else '-'
                
                # Get companies for filter (root only)
                companies = Company.objects.all() if request.user.role == 'root' else []
                
                context = {
                    'page_obj': page_obj,
                    'companies': companies,
                    'can_edit': can_edit_record(request.user),
                    'can_delete': can_delete_record(request.user),
                }
                return render(request, 'attendance_list.html', context)
            else:
                # No matching records, return empty queryset
                queryset = queryset.none()
                logger.info(f"Applied overstay > {hours} hours filter, found 0 records")
        except (ValueError, IndexError):
            logger.warning(f"Invalid overstay filter format: {overstay_filter}")
    
    # Order by date descending and ep_no
    queryset = queryset.order_by('-date', 'ep_no')
    
    logger.info(f"Filters applied, proceeding to pagination")
    
    # Pagination with 500 records per page for better performance
    paginator = Paginator(queryset.select_related('company'), 500)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Add shift_code attribute to each record
    for record in page_obj:
        if record.shift and '(' in str(record.shift):
            record.shift_code = str(record.shift).split('(')[0].strip()
        else:
            record.shift_code = str(record.shift) if record.shift else '-'
    
    # Get companies for filter (root only)
    companies = Company.objects.all() if request.user.role == 'root' else []
    
    context = {
        'page_obj': page_obj,
        'companies': companies,
        'can_edit': can_edit_record(request.user),
        'can_delete': can_delete_record(request.user),
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'attendance_list.html', context)


@login_required
@company_access_required
def attendance_export_view(request):
    """Export attendance records to XLSX with same filters as list view"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from .services.access_control_service import AccessControlService
    
    # Base queryset based on user role
    if request.user.role == 'root':
        queryset = AttendanceRecord.objects.all()
    elif request.user.role == 'user1':
        # User1: Filter by company and assigned employees
        queryset = AttendanceRecord.objects.filter(company=request.user.company)
        queryset = AccessControlService.filter_queryset_by_access(queryset, request.user)
    else:
        queryset = AttendanceRecord.objects.filter(company=request.user.company)
    
    # Apply same filters as list view
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    company_id = request.GET.get('company')
    ep_no = request.GET.get('ep_no')
    status = request.GET.get('status')
    overstay_filter = request.GET.get('overstay_filter')
    
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)
    if company_id and request.user.role == 'root':
        queryset = queryset.filter(company_id=company_id)
    if ep_no:
        queryset = queryset.filter(ep_no__icontains=ep_no)
    if status:
        queryset = queryset.filter(status=status)
    
    # Apply overstay filter
    if overstay_filter == 'has_overstay':
        queryset = queryset.exclude(Q(overstay='') | Q(overstay='00:00') | Q(overstay='-') | Q(overstay__isnull=True))
    elif overstay_filter == 'no_overstay':
        queryset = queryset.filter(Q(overstay='') | Q(overstay='00:00') | Q(overstay='-') | Q(overstay__isnull=True))
    elif overstay_filter and overstay_filter.startswith('range_'):
        # Filter for records with overstay in a specific range
        try:
            parts = overstay_filter.split('_')
            min_hours = int(parts[1])
            max_hours = int(parts[2])
            all_records = queryset.exclude(Q(overstay='') | Q(overstay='00:00') | Q(overstay='-') | Q(overstay__isnull=True))
            
            filtered_ids = []
            for record in all_records.iterator(chunk_size=500):
                try:
                    overstay_str = str(record.overstay).strip()
                    if ':' in overstay_str:
                        parts = overstay_str.split(':')
                        overstay_hours = int(parts[0])
                        overstay_minutes = int(parts[1]) if len(parts) > 1 else 0
                        total_hours = overstay_hours + (overstay_minutes / 60.0)
                        
                        if min_hours <= total_hours < max_hours:
                            filtered_ids.append(record.id)
                except (ValueError, AttributeError):
                    continue
            
            if filtered_ids:
                chunk_size = 900
                id_chunks = [filtered_ids[i:i + chunk_size] for i in range(0, len(filtered_ids), chunk_size)]
                
                q_objects = Q()
                for chunk in id_chunks:
                    q_objects |= Q(id__in=chunk)
                
                queryset = queryset.filter(q_objects)
            else:
                queryset = queryset.none()
        except (ValueError, IndexError):
            pass
    elif overstay_filter and overstay_filter.startswith('gt_'):
        # Filter for records with overstay greater than X hours
        try:
            hours = int(overstay_filter.split('_')[1])
            all_records = queryset.exclude(Q(overstay='') | Q(overstay='00:00') | Q(overstay='-') | Q(overstay__isnull=True))
            
            filtered_ids = []
            for record in all_records.iterator(chunk_size=500):
                try:
                    overstay_str = str(record.overstay).strip()
                    if ':' in overstay_str:
                        parts = overstay_str.split(':')
                        overstay_hours = int(parts[0])
                        overstay_minutes = int(parts[1]) if len(parts) > 1 else 0
                        total_hours = overstay_hours + (overstay_minutes / 60.0)
                        
                        if total_hours > hours:
                            filtered_ids.append(record.id)
                except (ValueError, AttributeError):
                    continue
            
            # Split IDs into chunks to avoid SQLite's 999 variable limit
            if filtered_ids:
                chunk_size = 900
                id_chunks = [filtered_ids[i:i + chunk_size] for i in range(0, len(filtered_ids), chunk_size)]
                
                q_objects = Q()
                for chunk in id_chunks:
                    q_objects |= Q(id__in=chunk)
                
                queryset = queryset.filter(q_objects)
            else:
                queryset = queryset.none()
        except (ValueError, IndexError):
            pass
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Records"
    
    # Header styling
    header_fill = PatternFill(start_color="4A70A9", end_color="4A70A9", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers in correct order
    headers = ['EP NO', 'EP NAME', 'COMPANY NAME', 'DATE', 'SHIFT', 'IN', 'OUT', 'IN (2)', 'OUT (2)', 'IN (3)', 'OUT (3)', 'HOURS', 'OVERSTAY', 'STATUS', 'OVERTIME', 'OVERTIME TO MANDAYS']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data in correct column order
    for row_num, record in enumerate(queryset.select_related('company'), 2):
        # Extract shift code
        shift_code = record.shift.split('(')[0].strip() if record.shift and '(' in record.shift else (record.shift or '')
        
        ws.cell(row=row_num, column=1, value=record.ep_no)
        ws.cell(row=row_num, column=2, value=record.ep_name)
        ws.cell(row=row_num, column=3, value=record.company.name if record.company else '')
        ws.cell(row=row_num, column=4, value=record.date.strftime('%d-%m-%Y') if record.date else '')
        ws.cell(row=row_num, column=5, value=shift_code)
        ws.cell(row=row_num, column=6, value=record.in_time.strftime('%H:%M') if record.in_time else '')
        ws.cell(row=row_num, column=7, value=record.out_time.strftime('%H:%M') if record.out_time else '')
        ws.cell(row=row_num, column=8, value=record.in_time_2.strftime('%H:%M') if record.in_time_2 else '')
        ws.cell(row=row_num, column=9, value=record.out_time_2.strftime('%H:%M') if record.out_time_2 else '')
        ws.cell(row=row_num, column=10, value=record.in_time_3.strftime('%H:%M') if record.in_time_3 else '')
        ws.cell(row=row_num, column=11, value=record.out_time_3.strftime('%H:%M') if record.out_time_3 else '')
        ws.cell(row=row_num, column=12, value=record.hours or '')
        ws.cell(row=row_num, column=13, value=record.overstay or '')
        ws.cell(row=row_num, column=14, value=record.status)
        ws.cell(row=row_num, column=15, value=record.overtime or '')
        ws.cell(row=row_num, column=16, value=record.overtime_to_mandays or '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="attendance_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response


@role_required(['root', 'admin'])
def attendance_edit_view(request, record_id):
    """Edit attendance record"""
    record = get_object_or_404(AttendanceRecord, id=record_id)
    
    # Check company access
    if not check_record_company_access(request.user, record):
        messages.error(request, 'You do not have permission to edit this record.')
        return HttpResponseForbidden('Access Denied')
    
    if request.method == 'POST':
        from .forms import AttendanceRecordForm
        form = AttendanceRecordForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            messages.success(request, 'Attendance record updated successfully.')
            return redirect('core:attendance_list')
    else:
        from .forms import AttendanceRecordForm
        form = AttendanceRecordForm(instance=record)
    
    context = {
        'form': form,
        'record': record,
    }
    return render(request, 'attendance_edit.html', context)


@role_required(['root', 'admin'])
def attendance_delete_view(request, record_id):
    """Delete attendance record"""
    record = get_object_or_404(AttendanceRecord, id=record_id)
    
    # Check company access
    if not check_record_company_access(request.user, record):
        messages.error(request, 'You do not have permission to delete this record.')
        return HttpResponseForbidden('Access Denied')
    
    if request.method == 'POST':
        record.delete()
        messages.success(request, 'Attendance record deleted successfully.')
        return redirect('core:attendance_list')
    
    return redirect('core:attendance_list')


@role_required(['root'])
def attendance_delete_all_view(request):
    """Delete all attendance records and companies (root only)"""
    if request.method == 'POST':
        attendance_count = AttendanceRecord.objects.all().count()
        company_count = Company.objects.all().count()
        
        # Delete all attendance records
        AttendanceRecord.objects.all().delete()
        
        # Delete all companies
        Company.objects.all().delete()
        
        logger.info(f'Root user {request.user.username} deleted all {attendance_count} attendance records and {company_count} companies')
        messages.success(request, f'Successfully deleted all {attendance_count} attendance records and {company_count} companies.')
        return redirect('core:attendance_list')
    
    return redirect('core:attendance_list')


@login_required
@role_required(['root', 'admin'])
def upload_progress_view(request):
    """API endpoint to check upload progress"""
    from django.http import JsonResponse
    import json
    import os
    
    progress_file = request.session.get('progress_file')
    if not progress_file or not os.path.exists(progress_file):
        return JsonResponse({'status': 'no_upload', 'processed': 0, 'total': 0, 'percentage': 0})
    
    try:
        with open(progress_file, 'r') as f:
            progress = json.load(f)
        return JsonResponse(progress)
    except Exception as e:
        return JsonResponse({'status': 'error', 'processed': 0, 'total': 0, 'percentage': 0, 'error': str(e)})


@login_required
def export_csv_view(request):
    """Export attendance records as CSV"""
    # Base queryset based on user role
    if request.user.role == 'root':
        queryset = AttendanceRecord.objects.all()
    else:
        queryset = AttendanceRecord.objects.filter(company=request.user.company)
    
    # Apply same filters as list view
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    company_id = request.GET.get('company')
    ep_no = request.GET.get('ep_no')
    status = request.GET.get('status')
    overstay_filter = request.GET.get('overstay_filter')
    
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)
    if company_id and request.user.role == 'root':
        queryset = queryset.filter(company_id=company_id)
    if ep_no:
        queryset = queryset.filter(ep_no__icontains=ep_no)
    if status:
        queryset = queryset.filter(status=status)
    
    # Apply overstay filter
    if overstay_filter == 'has_overstay':
        queryset = queryset.exclude(Q(overstay='') | Q(overstay='00:00') | Q(overstay='-') | Q(overstay__isnull=True))
    elif overstay_filter == 'no_overstay':
        queryset = queryset.filter(Q(overstay='') | Q(overstay='00:00') | Q(overstay='-') | Q(overstay__isnull=True))
    elif overstay_filter and overstay_filter.startswith('range_'):
        # Filter for records with overstay in a specific range
        try:
            parts = overstay_filter.split('_')
            min_hours = int(parts[1])
            max_hours = int(parts[2])
            all_records = queryset.exclude(Q(overstay='') | Q(overstay='00:00') | Q(overstay='-') | Q(overstay__isnull=True))
            
            filtered_ids = []
            for record in all_records.iterator(chunk_size=500):
                try:
                    overstay_str = str(record.overstay).strip()
                    if ':' in overstay_str:
                        parts = overstay_str.split(':')
                        overstay_hours = int(parts[0])
                        overstay_minutes = int(parts[1]) if len(parts) > 1 else 0
                        total_hours = overstay_hours + (overstay_minutes / 60.0)
                        
                        if min_hours <= total_hours < max_hours:
                            filtered_ids.append(record.id)
                except (ValueError, AttributeError):
                    continue
            
            if filtered_ids:
                chunk_size = 900
                id_chunks = [filtered_ids[i:i + chunk_size] for i in range(0, len(filtered_ids), chunk_size)]
                
                q_objects = Q()
                for chunk in id_chunks:
                    q_objects |= Q(id__in=chunk)
                
                queryset = queryset.filter(q_objects)
            else:
                queryset = queryset.none()
        except (ValueError, IndexError):
            pass
    elif overstay_filter and overstay_filter.startswith('gt_'):
        # Filter for records with overstay greater than X hours
        try:
            hours = int(overstay_filter.split('_')[1])
            all_records = queryset.exclude(Q(overstay='') | Q(overstay='00:00') | Q(overstay='-') | Q(overstay__isnull=True))
            
            filtered_ids = []
            for record in all_records.iterator(chunk_size=500):
                try:
                    overstay_str = str(record.overstay).strip()
                    if ':' in overstay_str:
                        parts = overstay_str.split(':')
                        overstay_hours = int(parts[0])
                        overstay_minutes = int(parts[1]) if len(parts) > 1 else 0
                        total_hours = overstay_hours + (overstay_minutes / 60.0)
                        
                        if total_hours > hours:
                            filtered_ids.append(record.id)
                except (ValueError, AttributeError):
                    continue
            
            # Split IDs into chunks to avoid SQLite's 999 variable limit
            if filtered_ids:
                chunk_size = 900
                id_chunks = [filtered_ids[i:i + chunk_size] for i in range(0, len(filtered_ids), chunk_size)]
                
                q_objects = Q()
                for chunk in id_chunks:
                    q_objects |= Q(id__in=chunk)
                
                queryset = queryset.filter(q_objects)
            else:
                queryset = queryset.none()
        except (ValueError, IndexError):
            pass
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="attendance_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    # Write header
    writer.writerow([
        'EP NO', 'EP NAME', 'COMPANY NAME', 'DATE', 'SHIFT', 'OVERSTAY', 'STATUS',
        'IN', 'OUT', 'IN (2)', 'OUT (2)', 'IN (3)', 'OUT (3)', 'OVERTIME', 'OVERTIME TO MANDAYS'
    ])
    
    # Write data
    for record in queryset.select_related('company'):
        writer.writerow([
            record.ep_no,
            record.ep_name,
            record.company.name,
            record.date.strftime('%Y-%m-%d'),
            record.shift,
            record.overstay,
            record.status,
            record.in_time.strftime('%H:%M') if record.in_time else '',
            record.out_time.strftime('%H:%M') if record.out_time else '',
            record.in_time_2.strftime('%H:%M') if record.in_time_2 else '',
            record.out_time_2.strftime('%H:%M') if record.out_time_2 else '',
            record.in_time_3.strftime('%H:%M') if record.in_time_3 else '',
            record.out_time_3.strftime('%H:%M') if record.out_time_3 else '',
            record.overtime.strftime('%H:%M') if record.overtime else '',
            record.overtime_to_mandays.strftime('%H:%M') if record.overtime_to_mandays else '',
        ])
    
    return response


@role_required(['root', 'admin'])
def download_csv_template(request):
    """Download empty CSV template with headers"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="attendance_template.csv"'
    
    writer = csv.writer(response)
    # Write header row with all required and optional columns in correct order
    writer.writerow([
        'EP NO', 'EP NAME', 'COMPANY NAME', 'DATE', 'SHIFT', 
        'IN', 'OUT', 'IN (2)', 'OUT (2)', 'IN (3)', 'OUT (3)', 
        'HOURS', 'OVERSTAY', 'STATUS', 'OVERTIME', 'OVERTIME TO MANDAYS'
    ])
    
    logger.info(f'CSV template downloaded by {request.user.username}')
    return response


@role_required(['root', 'admin'])
def upload_logs_view(request):
    """View upload logs"""
    if request.user.role == 'root':
        logs = UploadLog.objects.all()
    else:
        logs = UploadLog.objects.filter(user__company=request.user.company)
    
    # Pagination
    paginator = Paginator(logs.select_related('user'), 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
    }
    return render(request, 'upload_logs.html', context)


@role_required(['root', 'admin'])
def user_list_view(request):
    """List users"""
    if request.user.role == 'root':
        users = User.objects.all()
    else:
        users = User.objects.filter(company=request.user.company, role='user1')
    
    context = {
        'users': users.select_related('company'),
    }
    return render(request, 'user_list.html', context)


@role_required(['root', 'admin'])
def user_create_view(request):
    """Create new user"""
    if request.method == 'POST':
        from .forms import UserForm
        form = UserForm(request.POST, request_user=request.user)
        if form.is_valid():
            user = form.save(commit=False)
            
            # Auto-assign company for admin users
            if request.user.role == 'admin':
                user.company = request.user.company
                user.role = 'user1'
            
            # Set password
            password = form.cleaned_data.get('password')
            if password:
                user.set_password(password)
            
            user.save()
            messages.success(request, f'User {user.username} created successfully.')
            return redirect('core:user_list')
    else:
        from .forms import UserForm
        form = UserForm(request_user=request.user)
    
    context = {
        'form': form,
        'action': 'Create',
    }
    return render(request, 'user_form.html', context)


@role_required(['root', 'admin'])
def user_edit_view(request, user_id):
    """Edit user"""
    user = get_object_or_404(User, id=user_id)
    
    # Check permissions
    if request.user.role == 'admin':
        if user.role != 'user1' or user.company != request.user.company:
            messages.error(request, 'You do not have permission to edit this user.')
            return HttpResponseForbidden('Access Denied')
    
    if request.method == 'POST':
        from .forms import UserForm
        form = UserForm(request.POST, instance=user, request_user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, f'User {user.username} updated successfully.')
            return redirect('core:user_list')
    else:
        from .forms import UserForm
        form = UserForm(instance=user, request_user=request.user)
    
    context = {
        'form': form,
        'action': 'Edit',
        'user_obj': user,
    }
    return render(request, 'user_form.html', context)


@role_required(['root'])
def user_delete_view(request, user_id):
    """Delete user (root only)"""
    user = get_object_or_404(User, id=user_id)
    
    # Prevent deleting yourself
    if user.id == request.user.id:
        messages.error(request, 'You cannot delete your own account.')
        return redirect('core:user_list')
    
    # Prevent deleting root users
    if user.role == 'root':
        messages.error(request, 'Root users cannot be deleted.')
        return redirect('core:user_list')
    
    if request.method == 'POST':
        username = user.username
        user.delete()
        logger.info(f'User deleted by {request.user.username}: {username}')
        messages.success(request, f'User {username} has been deleted successfully.')
        return redirect('core:user_list')
    
    context = {
        'user_obj': user,
    }
    return render(request, 'user_delete_confirm.html', context)


# Error handlers
def handler404(request, exception):
    """Custom 404 error handler"""
    return render(request, '404.html', status=404)


def handler403(request, exception):
    """Custom 403 error handler"""
    return render(request, '403.html', status=403)


def handler500(request):
    """Custom 500 error handler"""
    return render(request, '500.html', status=500)


# Backup and Restore Views

@role_required(['root'])
def backup_data_view(request):
    """Create and download backup (root only)"""
    from .models import BackupLog
    from .services.backup_service import BackupService
    import json
    
    if request.method == 'POST':
        backup_type = request.POST.get('backup_type', 'full')
        since_date_str = request.POST.get('since_date', '')
        
        since_date = None
        if backup_type == 'incremental' and since_date_str:
            try:
                since_date = datetime.strptime(since_date_str, '%Y-%m-%d')
            except ValueError:
                messages.error(request, 'Invalid date format')
                return redirect('core:backup_data')
        
        # Create backup
        service = BackupService()
        if backup_type == 'incremental' and since_date:
            result = service.create_incremental_backup(since_date)
        else:
            result = service.create_backup(backup_type='full')
        
        if result['success']:
            # Create backup log
            operation = 'backup_incremental' if backup_type == 'incremental' else 'backup_full'
            backup_log = BackupLog.objects.create(
                user=request.user,
                operation=operation,
                filename=f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                companies_count=result['companies_count'],
                records_count=result['records_count'],
                success=True
            )
            
            # Store backup data in session for download
            request.session['backup_data'] = json.dumps(result['data'])
            request.session['backup_filename'] = backup_log.filename
            
            messages.success(request, f'Backup created successfully! {result["companies_count"]} companies, {result["records_count"]} records.')
            return redirect('core:download_backup')
        else:
            # Log failure
            BackupLog.objects.create(
                user=request.user,
                operation='backup_full',
                filename='failed_backup.json',
                success=False,
                error_message=result.get('error', 'Unknown error')
            )
            messages.error(request, f'Backup failed: {result.get("error", "Unknown error")}')
    
    # Get last backup and recent backups
    last_backup = BackupLog.objects.filter(
        operation__in=['backup_full', 'backup_incremental']
    ).first()
    
    recent_backups = BackupLog.objects.filter(
        operation__in=['backup_full', 'backup_incremental']
    )[:10]
    
    context = {
        'last_backup': last_backup,
        'recent_backups': recent_backups
    }
    return render(request, 'backup_data.html', context)


@role_required(['root'])
def download_backup_view(request):
    """Download the backup file"""
    import json
    
    backup_data_json = request.session.get('backup_data')
    backup_filename = request.session.get('backup_filename', 'backup.json')
    
    if not backup_data_json:
        messages.error(request, 'No backup data found. Please create a backup first.')
        return redirect('core:backup_data')
    
    # Clear session data
    del request.session['backup_data']
    del request.session['backup_filename']
    
    # Create response
    response = HttpResponse(backup_data_json, content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename="{backup_filename}"'
    
    return response


@role_required(['root'])
def restore_data_view(request):
    """Restore data from backup file (root only)"""
    from .models import BackupLog
    
    # Get recent restore operations
    recent_restores = BackupLog.objects.filter(operation='restore')[:10]
    
    context = {
        'recent_restores': recent_restores
    }
    return render(request, 'restore_data.html', context)


@role_required(['root'])
def restore_preview_view(request):
    """Preview restore changes before applying"""
    import json
    from .services.restore_service import RestoreService
    
    if request.method != 'POST':
        return redirect('core:restore_data')
    
    backup_file = request.FILES.get('backup_file')
    merge_strategy = request.POST.get('merge_strategy', 'backup_wins')
    
    if not backup_file:
        messages.error(request, 'Please select a backup file')
        return redirect('core:restore_data')
    
    try:
        # Read and parse backup file
        backup_data = json.load(backup_file)
        
        # Validate backup
        service = RestoreService()
        validation = service.validate_backup(backup_data)
        
        if not validation['valid']:
            for error in validation['errors']:
                messages.error(request, f'Validation error: {error}')
            return redirect('core:restore_data')
        
        # Preview changes
        preview = service.preview_changes(backup_data)
        
        # Store backup data in session for apply step
        request.session['restore_backup_data'] = json.dumps(backup_data)
        request.session['restore_merge_strategy'] = merge_strategy
        request.session['restore_filename'] = backup_file.name
        
        context = {
            'preview': preview,
            'merge_strategy': merge_strategy,
            'filename': backup_file.name,
            'total_records': len(backup_data.get('attendance_records', []))
        }
        return render(request, 'restore_preview.html', context)
        
    except json.JSONDecodeError:
        messages.error(request, 'Invalid JSON file')
        return redirect('core:restore_data')
    except Exception as e:
        messages.error(request, f'Error reading backup file: {str(e)}')
        return redirect('core:restore_data')


@role_required(['root'])
def restore_apply_view(request):
    """Apply the restore operation"""
    import json
    from .services.restore_service import RestoreService
    from .models import BackupLog
    
    if request.method != 'POST':
        return redirect('core:restore_data')
    
    # Get backup data from session
    backup_data_json = request.session.get('restore_backup_data')
    merge_strategy = request.session.get('restore_merge_strategy', 'backup_wins')
    filename = request.session.get('restore_filename', 'unknown.json')
    
    if not backup_data_json:
        messages.error(request, 'No backup data found. Please upload a backup file first.')
        return redirect('core:restore_data')
    
    try:
        backup_data = json.loads(backup_data_json)
        
        # Perform restore
        service = RestoreService()
        result = service.restore_backup(backup_data, merge_strategy=merge_strategy)
        
        # Create backup log
        backup_log = BackupLog.objects.create(
            user=request.user,
            operation='restore',
            filename=filename,
            companies_count=len(backup_data.get('companies', [])),
            records_count=len(backup_data.get('attendance_records', [])),
            records_added=result.get('added', 0),
            records_updated=result.get('updated', 0),
            records_skipped=result.get('skipped', 0),
            success=result.get('success', False),
            error_message='\n'.join(result.get('errors', []))
        )
        
        # Clear session data
        if 'restore_backup_data' in request.session:
            del request.session['restore_backup_data']
        if 'restore_merge_strategy' in request.session:
            del request.session['restore_merge_strategy']
        if 'restore_filename' in request.session:
            del request.session['restore_filename']
        
        if result['success']:
            messages.success(
                request,
                f'Restore completed successfully! Added: {result["added"]}, Updated: {result["updated"]}, Skipped: {result["skipped"]}'
            )
        else:
            messages.error(request, f'Restore failed: {", ".join(result.get("errors", ["Unknown error"]))}')
        
        return redirect('core:restore_data')
        
    except Exception as e:
        messages.error(request, f'Error during restore: {str(e)}')
        return redirect('core:restore_data')



@login_required
@role_required(['user1'])
def request_access_view(request):
    """User1 request access to employees"""
    from .services.request_approval_service import RequestApprovalService
    
    if request.method == 'POST':
        ep_nos_input = request.POST.get('ep_nos', '').strip()
        access_type = request.POST.get('access_type', 'permanent')
        access_from = request.POST.get('access_from', '').strip()
        access_to = request.POST.get('access_to', '').strip()
        justification = request.POST.get('justification', '').strip()
        
        # Validate inputs
        if not ep_nos_input:
            messages.error(request, 'Please enter at least one Employee Number.')
            return render(request, 'request_access.html')
        
        if not justification:
            messages.error(request, 'Please provide a justification for your request.')
            return render(request, 'request_access.html')
        
        # Parse EP NOs
        valid_ep_nos, invalid_ep_nos = RequestApprovalService.parse_bulk_ep_nos(ep_nos_input)
        
        if invalid_ep_nos:
            messages.warning(request, f'Skipped {len(invalid_ep_nos)} invalid/empty EP NOs.')
        
        if not valid_ep_nos:
            messages.error(request, 'No valid Employee Numbers provided.')
            return render(request, 'request_access.html')
        
        # Prepare dates
        dates = {}
        if access_type == 'date_range':
            if not access_from or not access_to:
                messages.error(request, 'Date range access requires both start and end dates.')
                return render(request, 'request_access.html', {
                    'ep_nos': ep_nos_input,
                    'access_type': access_type,
                    'justification': justification
                })
            dates = {
                'access_from': access_from,
                'access_to': access_to
            }
        
        # Create requests
        try:
            requests = RequestApprovalService.create_request(
                user=request.user,
                ep_nos=valid_ep_nos,
                access_type=access_type,
                dates=dates,
                justification=justification
            )
            
            messages.success(
                request,
                f'Successfully submitted {len(requests)} access request(s). '
                f'You will be notified when they are reviewed.'
            )
            return redirect('core:my_requests')
        except Exception as e:
            logger.error(f'Error creating access request: {e}')
            messages.error(request, f'Error creating request: {str(e)}')
    
    return render(request, 'request_access.html')


@login_required
@role_required(['user1'])
def my_requests_view(request):
    """User1 view their access requests"""
    from .models import AccessRequest
    
    # Get all requests for current user, sorted by most recent first
    requests = AccessRequest.objects.filter(
        requester=request.user
    ).order_by('-created_at')
    
    # Pagination
    paginator = Paginator(requests, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
    }
    return render(request, 'my_requests.html', context)


@login_required
@role_required(['user1'])
def cancel_request_view(request, request_id):
    """User1 cancel their pending request"""
    from .services.request_approval_service import RequestApprovalService
    
    if request.method == 'POST':
        try:
            RequestApprovalService.cancel_request(request_id, request.user)
            messages.success(request, 'Request cancelled successfully.')
        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            logger.error(f'Error cancelling request: {e}')
            messages.error(request, 'Error cancelling request.')
    
    return redirect('core:my_requests')



@login_required
@role_required(['admin', 'root'])
def approve_requests_view(request):
    """Admin view and approve/reject access requests"""
    from .models import AccessRequest
    
    # Get all pending requests
    pending_requests = AccessRequest.objects.filter(
        status='pending'
    ).select_related('requester', 'company').order_by('-created_at')
    
    # Pagination
    paginator = Paginator(pending_requests, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
    }
    return render(request, 'approve_requests.html', context)


@login_required
@role_required(['admin', 'root'])
def approve_request_action(request, request_id):
    """Admin approve a request"""
    from .services.request_approval_service import RequestApprovalService
    
    if request.method == 'POST':
        try:
            RequestApprovalService.approve_request(request_id, request.user)
            messages.success(request, 'Request approved successfully.')
        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            logger.error(f'Error approving request: {e}')
            messages.error(request, 'Error approving request.')
    
    return redirect('core:approve_requests')


@login_required
@role_required(['admin', 'root'])
def reject_request_action(request, request_id):
    """Admin reject a request"""
    from .services.request_approval_service import RequestApprovalService
    
    if request.method == 'POST':
        reason = request.POST.get('reason', '').strip()
        
        try:
            RequestApprovalService.reject_request(request_id, request.user, reason)
            messages.success(request, 'Request rejected.')
        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            logger.error(f'Error rejecting request: {e}')
            messages.error(request, 'Error rejecting request.')
    
    return redirect('core:approve_requests')



@login_required
@role_required(['admin', 'root'])
def manage_assignments_view(request):
    """Admin manage employee assignments"""
    from .models import EmployeeAssignment
    
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        
        # Validate file extension
        if not csv_file.name.lower().endswith('.csv'):
            messages.error(request, 'Please upload a valid CSV file.')
            return redirect('core:manage_assignments')
        
        try:
            import csv
            import io
            
            # Read CSV
            file_data = csv_file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(file_data))
            
            created_count = 0
            error_count = 0
            errors = []
            
            # Expected columns: username, ep_no, access_from, access_to
            for row_num, row in enumerate(csv_reader, start=2):
                try:
                    username = row.get('username', '').strip()
                    ep_no = row.get('ep_no', '').strip()
                    access_from = row.get('access_from', '').strip()
                    access_to = row.get('access_to', '').strip()
                    
                    if not username or not ep_no:
                        errors.append(f"Row {row_num}: Missing username or ep_no")
                        error_count += 1
                        continue
                    
                    # Find user
                    try:
                        user = User.objects.get(username=username, role='user1')
                    except User.DoesNotExist:
                        errors.append(f"Row {row_num}: User '{username}' not found or not User1")
                        error_count += 1
                        continue
                    
                    # Parse dates
                    from datetime import datetime
                    access_from_date = None
                    access_to_date = None
                    
                    if access_from:
                        try:
                            access_from_date = datetime.strptime(access_from, '%Y-%m-%d').date()
                        except ValueError:
                            errors.append(f"Row {row_num}: Invalid access_from date format (use YYYY-MM-DD)")
                            error_count += 1
                            continue
                    
                    if access_to:
                        try:
                            access_to_date = datetime.strptime(access_to, '%Y-%m-%d').date()
                        except ValueError:
                            errors.append(f"Row {row_num}: Invalid access_to date format (use YYYY-MM-DD)")
                            error_count += 1
                            continue
                    
                    # Create assignment
                    EmployeeAssignment.objects.create(
                        user=user,
                        ep_no=ep_no,
                        ep_name=f"Employee {ep_no}",
                        company=user.company,
                        access_from=access_from_date,
                        access_to=access_to_date,
                        assigned_by=request.user,
                        source='admin',
                        is_active=True
                    )
                    
                    # Log assignment
                    from .models import AccessRequestAuditLog
                    AccessRequestAuditLog.create_log_entry(
                        action='assignment_created',
                        actor=request.user,
                        target_user=user,
                        target_ep_no=ep_no,
                        details={
                            'source': 'csv_upload'
                        }
                    )
                    
                    created_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    error_count += 1
            
            # Show results
            if created_count > 0:
                messages.success(request, f'Successfully created {created_count} assignment(s).')
            
            if error_count > 0:
                messages.warning(request, f'{error_count} error(s) occurred.')
                for error in errors[:5]:  # Show first 5 errors
                    messages.error(request, error)
                if len(errors) > 5:
                    messages.info(request, f'... and {len(errors) - 5} more errors.')
        
        except Exception as e:
            logger.error(f'Error processing assignment CSV: {e}')
            messages.error(request, f'Error processing CSV: {str(e)}')
        
        return redirect('core:manage_assignments')
    
    # Get all assignments
    assignments = EmployeeAssignment.objects.select_related(
        'user', 'company', 'assigned_by'
    ).order_by('-assigned_at')
    
    # Pagination
    paginator = Paginator(assignments, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
    }
    return render(request, 'manage_assignments.html', context)


@login_required
@role_required(['admin', 'root'])
def remove_assignment_view(request, assignment_id):
    """Admin remove an assignment"""
    from .models import EmployeeAssignment, AccessRequestAuditLog
    
    if request.method == 'POST':
        try:
            assignment = EmployeeAssignment.objects.get(id=assignment_id)
            
            # Mark as inactive
            assignment.is_active = False
            assignment.save()
            
            # Log removal
            AccessRequestAuditLog.create_log_entry(
                action='assignment_removed',
                actor=request.user,
                target_user=assignment.user,
                target_ep_no=assignment.ep_no,
                details={
                    'assignment_id': assignment.id
                }
            )
            
            messages.success(request, f'Assignment removed: {assignment.user.username} → {assignment.ep_no}')
        except EmployeeAssignment.DoesNotExist:
            messages.error(request, 'Assignment not found.')
        except Exception as e:
            logger.error(f'Error removing assignment: {e}')
            messages.error(request, 'Error removing assignment.')
    
    return redirect('core:manage_assignments')



# Manday Summary Views

@role_required(['root', 'admin'])
def upload_mandays_view(request):
    """Handle manday CSV/Excel file upload and processing"""
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        
        # Validate file extension
        allowed_extensions = ['.csv', '.xls', '.xlsx']
        if not any(csv_file.name.lower().endswith(ext) for ext in allowed_extensions):
            logger.warning(f'Invalid manday file upload attempt by {request.user.username}: {csv_file.name}')
            messages.error(request, 'Please upload a valid CSV, XLS, or XLSX file.')
            return redirect('core:upload_mandays')
        
        logger.info(f'Manday CSV upload started by {request.user.username}: {csv_file.name}')
        
        # Create a unique upload ID for this session
        from django.core.cache import cache
        upload_id = f"manday_upload_{request.user.id}_{int(timezone.now().timestamp())}"
        
        # Initialize progress in cache
        cache.set(upload_id, {'processed': 0, 'total': 0, 'status': 'starting'}, timeout=3600)
        
        # Store upload_id in session
        request.session['current_manday_upload_id'] = upload_id
        
        # Process CSV with progress callback
        from .manday_processor import MandayProcessor
        from .models import MandayUploadLog
        processor = MandayProcessor()
        
        # Set progress callback
        def progress_callback(processed, total):
            cache.set(upload_id, {
                'processed': processed,
                'total': total,
                'status': 'processing',
                'percentage': int((processed / total) * 100) if total > 0 else 0
            }, timeout=3600)
        
        processor.progress_callback = progress_callback
        result = processor.process_csv(csv_file, request.user)
        
        # Mark as complete
        cache.set(upload_id, {
            'processed': result.get('processed_rows', 0),
            'total': result.get('total_rows', 0),
            'status': 'complete',
            'percentage': 100
        }, timeout=3600)
        
        # Create upload log
        error_messages = '\n'.join(result['errors']) if result['errors'] else ''
        upload_log = MandayUploadLog.objects.create(
            user=request.user,
            filename=csv_file.name,
            success_count=result['success_count'],
            updated_count=result['updated_count'],
            error_count=result['error_count'],
            error_messages=error_messages
        )
        
        # Log results
        logger.info(
            f'Manday CSV processing completed by {request.user.username}: '
            f'Created={result["success_count"]}, Updated={result["updated_count"]}, Errors={result["error_count"]}'
        )
        
        if result['error_count'] > 0:
            logger.error(f'Manday CSV processing errors for {csv_file.name}: {error_messages[:200]}...')
        
        # Display results
        if result['success']:
            messages.success(
                request,
                f'Manday data processed successfully! Created: {result["success_count"]}, Updated: {result["updated_count"]}'
            )
        else:
            messages.warning(
                request,
                f'Manday data processed with errors. Created: {result["success_count"]}, Updated: {result["updated_count"]}, Errors: {result["error_count"]}'
            )
            if result['errors']:
                for error in result['errors'][:5]:  # Show first 5 errors
                    messages.error(request, error)
                if len(result['errors']) > 5:
                    messages.info(request, f'... and {len(result["errors"]) - 5} more errors. Check upload logs for details.')
        
        return redirect('core:upload_mandays')
    
    # Get recent upload logs for current user
    from .models import MandayUploadLog
    if request.user.role == 'root':
        recent_logs = MandayUploadLog.objects.all()[:10]
    else:
        recent_logs = MandayUploadLog.objects.filter(user__company=request.user.company)[:10]
    
    context = {
        'recent_logs': recent_logs
    }
    return render(request, 'upload_mandays.html', context)


@login_required
@company_access_required
def mandays_list_view(request):
    """List manday records with filtering and pagination"""
    from .services.access_control_service import AccessControlService
    from .models import MandaySummaryRecord
    
    # Base queryset based on user role
    if request.user.role == 'root':
        queryset = MandaySummaryRecord.objects.all()
    elif request.user.role == 'user1':
        # User1: Filter by company and assigned employees
        queryset = MandaySummaryRecord.objects.filter(company=request.user.company)
        # Note: For mandays, we don't filter by employee assignments like attendance
        # This is a business decision - adjust if needed
    else:
        # Admin: Filter by company only
        queryset = MandaySummaryRecord.objects.filter(company=request.user.company)
    
    # Apply filters
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    company_id = request.GET.get('company', '').strip()
    ep_no = request.GET.get('ep_no', '').strip()
    
    logger.info(f"Manday Filters - EP NO: '{ep_no}', Date From: '{date_from}', Date To: '{date_to}'")
    
    # Only apply date filters if values are provided
    if date_from:
        queryset = queryset.filter(punch_date__gte=date_from)
        logger.info(f"Applied date_from filter: {date_from}")
    if date_to:
        queryset = queryset.filter(punch_date__lte=date_to)
        logger.info(f"Applied date_to filter: {date_to}")
    if company_id and request.user.role == 'root':
        queryset = queryset.filter(company_id=company_id)
    if ep_no:
        queryset = queryset.filter(ep_no__icontains=ep_no)
        logger.info(f"Applied ep_no filter: {ep_no}")
    
    # Order by date descending and ep_no
    queryset = queryset.order_by('-punch_date', 'ep_no')
    
    logger.info(f"Filters applied, proceeding to pagination")
    
    # Pagination
    paginator = Paginator(queryset.select_related('company'), 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get companies for filter (root only)
    companies = Company.objects.all() if request.user.role == 'root' else []
    
    context = {
        'page_obj': page_obj,
        'companies': companies,
    }
    return render(request, 'mandays_list.html', context)


@login_required
@company_access_required
def mandays_export_view(request):
    """Export manday records to XLSX with same filters as list view"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from .models import MandaySummaryRecord
    
    # Base queryset based on user role
    if request.user.role == 'root':
        queryset = MandaySummaryRecord.objects.all()
    elif request.user.role == 'user1':
        queryset = MandaySummaryRecord.objects.filter(company=request.user.company)
    else:
        queryset = MandaySummaryRecord.objects.filter(company=request.user.company)
    
    # Apply same filters as list view
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    company_id = request.GET.get('company')
    ep_no = request.GET.get('ep_no')
    
    if date_from:
        queryset = queryset.filter(punch_date__gte=date_from)
    if date_to:
        queryset = queryset.filter(punch_date__lte=date_to)
    if company_id and request.user.role == 'root':
        queryset = queryset.filter(company_id=company_id)
    if ep_no:
        queryset = queryset.filter(ep_no__icontains=ep_no)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Manday Summary"
    
    # Header styling
    header_fill = PatternFill(start_color="4A70A9", end_color="4A70A9", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers in correct order
    headers = ['EP NO', 'PUNCH DATE', 'COMPANY', 'MANDAYS', 'REGULAR MANDAY HR', 'OT', 'TRADE', 'CONTRACT', 'PLANT', 'PLANT DESC']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data in correct column order
    for row_num, record in enumerate(queryset.select_related('company'), 2):
        ws.cell(row=row_num, column=1, value=record.ep_no)
        ws.cell(row=row_num, column=2, value=record.punch_date.strftime('%d-%m-%Y') if record.punch_date else '')
        ws.cell(row=row_num, column=3, value=record.company.name if record.company else '')
        ws.cell(row=row_num, column=4, value=float(record.mandays) if record.mandays else 0)
        ws.cell(row=row_num, column=5, value=float(record.regular_manday_hr) if record.regular_manday_hr else 0)
        ws.cell(row=row_num, column=6, value=float(record.ot) if record.ot else 0)
        ws.cell(row=row_num, column=7, value=record.trade or '')
        ws.cell(row=row_num, column=8, value=record.contract or '')
        ws.cell(row=row_num, column=9, value=record.plant or '')
        ws.cell(row=row_num, column=10, value=record.plant_desc or '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="mandays_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response


@role_required(['root', 'admin'])
def download_mandays_template(request):
    """Download empty CSV template with headers for manday upload"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="mandays_template.csv"'
    
    writer = csv.writer(response)
    # Write header row with all required and optional columns
    writer.writerow([
        'epNo', 'punchDate', 'mandays', 'regularMandayHr', 'ot',
        'trade', 'skill', 'contract', 'plant', 'plantDesc'
    ])
    
    logger.info(f'Manday CSV template downloaded by {request.user.username}')
    return response



# ============================================================================
# Excel File Upload Integration Views
# ============================================================================

@login_required
def excel_upload_view(request):
    """Excel file upload interface"""
    return render(request, 'excel_upload.html')


@login_required
def excel_dashboard_view(request):
    """Excel data dashboard"""
    return render(request, 'excel_dashboard.html')


@login_required
def excel_search_view(request):
    """Search and filter Excel data"""
    return render(request, 'excel_search.html')


@login_required
def excel_import_history_view(request):
    """View import history"""
    return render(request, 'excel_import_history.html')


@login_required
@user_passes_test(lambda u: u.role in ['root', 'admin'])
def excel_permissions_view(request):
    """Manage upload permissions (admin only)"""
    return render(request, 'excel_permissions.html')



@login_required
@company_access_required
def add_remark_view(request, record_id):
    """Add remark to attendance record (User1 only)"""
    # Only User1 can add remarks
    if request.user.role != 'user1':
        messages.error(request, 'Only User1 can add remarks to attendance records.')
        return redirect('core:attendance_list')
    
    # Get the attendance record
    record = get_object_or_404(AttendanceRecord, id=record_id)
    
    # Check if user has access to this record's company
    if request.user.company != record.company:
        messages.error(request, 'You do not have access to this record.')
        return redirect('core:attendance_list')
    
    # Check date range access for User1
    if request.user.assigned_date_from and record.date < request.user.assigned_date_from:
        messages.error(request, 'This record is outside your assigned date range.')
        return redirect('core:attendance_list')
    
    if request.user.assigned_date_to and record.date > request.user.assigned_date_to:
        messages.error(request, 'This record is outside your assigned date range.')
        return redirect('core:attendance_list')
    
    if request.method == 'POST':
        reason_id = request.POST.get('reason')
        remarks_text = request.POST.get('remarks_text', '').strip()
        
        if not reason_id:
            messages.error(request, 'Please select a reason.')
            return redirect('core:add_remark', record_id=record_id)
        
        if not remarks_text:
            messages.error(request, 'Please enter remarks.')
            return redirect('core:add_remark', record_id=record_id)
        
        try:
            reason = RemarkReason.objects.get(id=reason_id, company=request.user.company, is_active=True)
            
            # Create the remark
            AttendanceRemark.objects.create(
                attendance_record=record,
                ep_no=record.ep_no,
                date=record.date,
                reason=reason,
                remarks_text=remarks_text,
                created_by=request.user
            )
            
            messages.success(request, f'Remark added successfully for {record.ep_no} on {record.date}.')
            logger.info(f"User {request.user.username} added remark for {record.ep_no} on {record.date}")
            return redirect('core:attendance_list')
            
        except RemarkReason.DoesNotExist:
            messages.error(request, 'Invalid reason selected.')
            return redirect('core:add_remark', record_id=record_id)
    
    # GET request - show form
    reasons = RemarkReason.objects.filter(company=request.user.company, is_active=True)
    
    context = {
        'record': record,
        'reasons': reasons,
    }
    return render(request, 'add_remark.html', context)


@login_required
@company_access_required
def remarks_log_view(request):
    """View all remarks (Admin and Root only)"""
    if request.user.role not in ['admin', 'root']:
        messages.error(request, 'You do not have permission to view remarks log.')
        return redirect('core:dashboard')
    
    # Filter remarks based on user role
    if request.user.role == 'root':
        remarks = AttendanceRemark.objects.all().select_related(
            'attendance_record', 'reason', 'created_by', 'responded_by'
        )
    else:  # admin
        remarks = AttendanceRemark.objects.filter(
            attendance_record__company=request.user.company
        ).select_related(
            'attendance_record', 'reason', 'created_by', 'responded_by'
        )
    
    # Filters
    status_filter = request.GET.get('status', '')
    ep_no_filter = request.GET.get('ep_no', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if status_filter:
        remarks = remarks.filter(status=status_filter)
    
    if ep_no_filter:
        remarks = remarks.filter(ep_no__icontains=ep_no_filter)
    
    if date_from:
        remarks = remarks.filter(date__gte=date_from)
    
    if date_to:
        remarks = remarks.filter(date__lte=date_to)
    
    # Order by most recent
    remarks = remarks.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(remarks, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
    }
    return render(request, 'remarks_log.html', context)


@login_required
@role_required(['admin', 'root'])
def manage_remark_reasons_view(request):
    """Manage remark reasons/categories (Admin and Root only)"""
    if request.user.role == 'root':
        reasons = RemarkReason.objects.all().select_related('company', 'created_by')
    else:  # admin
        reasons = RemarkReason.objects.filter(company=request.user.company).select_related('created_by')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            reason_text = request.POST.get('reason', '').strip()
            if reason_text:
                RemarkReason.objects.create(
                    company=request.user.company,
                    reason=reason_text,
                    created_by=request.user
                )
                messages.success(request, f'Reason "{reason_text}" added successfully.')
                logger.info(f"User {request.user.username} added remark reason: {reason_text}")
            else:
                messages.error(request, 'Please enter a reason.')
        
        elif action == 'toggle':
            reason_id = request.POST.get('reason_id')
            try:
                reason = RemarkReason.objects.get(id=reason_id)
                if request.user.role == 'admin' and reason.company != request.user.company:
                    messages.error(request, 'You do not have permission to modify this reason.')
                else:
                    reason.is_active = not reason.is_active
                    reason.save()
                    status = 'activated' if reason.is_active else 'deactivated'
                    messages.success(request, f'Reason "{reason.reason}" {status}.')
            except RemarkReason.DoesNotExist:
                messages.error(request, 'Reason not found.')
        
        return redirect('core:manage_remark_reasons')
    
    context = {
        'reasons': reasons,
    }
    return render(request, 'manage_remark_reasons.html', context)



@login_required
@role_required(['admin', 'root'])
def export_remarks_log_view(request):
    """Export remarks log to Excel (Admin and Root only)"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    # Filter remarks based on user role
    if request.user.role == 'root':
        remarks = AttendanceRemark.objects.all().select_related(
            'attendance_record', 'reason', 'created_by', 'responded_by', 'attendance_record__company'
        )
    else:  # admin
        remarks = AttendanceRemark.objects.filter(
            attendance_record__company=request.user.company
        ).select_related(
            'attendance_record', 'reason', 'created_by', 'responded_by', 'attendance_record__company'
        )
    
    # Apply filters from request
    status_filter = request.GET.get('status', '')
    ep_no_filter = request.GET.get('ep_no', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if status_filter:
        remarks = remarks.filter(status=status_filter)
    if ep_no_filter:
        remarks = remarks.filter(ep_no__icontains=ep_no_filter)
    if date_from:
        remarks = remarks.filter(date__gte=date_from)
    if date_to:
        remarks = remarks.filter(date__lte=date_to)
    
    remarks = remarks.order_by('-created_at')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Remarks Log"
    
    # Define styles
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['EP NO', 'Employee Name', 'Date', 'Company', 'Reason', 'Remarks', 
               'Status', 'Submitted By', 'Submitted Date', 'Submitted Time', 
               'Admin Response', 'Responded By', 'Responded Date']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for row_num, remark in enumerate(remarks, 2):
        ws.cell(row=row_num, column=1).value = remark.ep_no
        ws.cell(row=row_num, column=2).value = remark.attendance_record.ep_name if remark.attendance_record else ''
        ws.cell(row=row_num, column=3).value = remark.date.strftime('%d/%m/%Y') if remark.date else ''
        ws.cell(row=row_num, column=4).value = remark.attendance_record.company.name if remark.attendance_record and remark.attendance_record.company else ''
        ws.cell(row=row_num, column=5).value = remark.reason.reason if remark.reason else ''
        ws.cell(row=row_num, column=6).value = remark.remarks_text
        ws.cell(row=row_num, column=7).value = remark.get_status_display()
        ws.cell(row=row_num, column=8).value = remark.created_by.username if remark.created_by else ''
        ws.cell(row=row_num, column=9).value = remark.created_at.strftime('%d/%m/%Y') if remark.created_at else ''
        ws.cell(row=row_num, column=10).value = remark.created_at.strftime('%H:%M:%S') if remark.created_at else ''
        ws.cell(row=row_num, column=11).value = remark.admin_response or ''
        ws.cell(row=row_num, column=12).value = remark.responded_by.username if remark.responded_by else ''
        ws.cell(row=row_num, column=13).value = remark.responded_at.strftime('%d/%m/%Y %H:%M') if remark.responded_at else ''
        
        # Apply borders
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_num).border = border
    
    # Adjust column widths
    column_widths = [12, 25, 12, 25, 20, 40, 12, 15, 15, 12, 40, 15, 18]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'remarks_log_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    logger.info(f"User {request.user.username} exported remarks log with {remarks.count()} records")
    return response


@login_required
@role_required(['root'])
def upload_remarks_log_view(request):
    """Upload remarks log from Excel (Root only)"""
    if request.method == 'POST':
        if 'file' not in request.FILES:
            messages.error(request, 'No file uploaded.')
            return redirect('core:upload_remarks_log')
        
        file = request.FILES['file']
        
        if not file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Please upload an Excel file (.xlsx or .xls)')
            return redirect('core:upload_remarks_log')
        
        try:
            import openpyxl
            from datetime import datetime
            
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            success_count = 0
            error_count = 0
            errors = []
            
            # Skip header row
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    ep_no = str(row[0]).strip() if row[0] else None
                    date_str = row[2]
                    company_name = str(row[3]).strip() if row[3] else None
                    reason_text = str(row[4]).strip() if row[4] else None
                    remarks_text = str(row[5]).strip() if row[5] else None
                    status = str(row[6]).strip().lower() if row[6] else 'pending'
                    created_by_username = str(row[7]).strip() if row[7] else None
                    admin_response = str(row[10]).strip() if row[10] else ''
                    
                    if not all([ep_no, date_str, company_name, reason_text, remarks_text]):
                        errors.append(f"Row {row_num}: Missing required fields")
                        error_count += 1
                        continue
                    
                    # Parse date
                    if isinstance(date_str, datetime):
                        date = date_str.date()
                    else:
                        date = datetime.strptime(str(date_str), '%d/%m/%Y').date()
                    
                    # Get company
                    company = Company.objects.filter(name=company_name).first()
                    if not company:
                        errors.append(f"Row {row_num}: Company '{company_name}' not found")
                        error_count += 1
                        continue
                    
                    # Get attendance record
                    attendance_record = AttendanceRecord.objects.filter(
                        ep_no=ep_no,
                        date=date,
                        company=company
                    ).first()
                    
                    if not attendance_record:
                        errors.append(f"Row {row_num}: Attendance record not found for {ep_no} on {date}")
                        error_count += 1
                        continue
                    
                    # Get or create reason
                    reason, _ = RemarkReason.objects.get_or_create(
                        company=company,
                        reason=reason_text,
                        defaults={'created_by': request.user, 'is_active': True}
                    )
                    
                    # Get created_by user
                    created_by = User.objects.filter(username=created_by_username).first() if created_by_username else request.user
                    
                    # Map status
                    status_map = {
                        'pending': 'pending',
                        'reviewed': 'reviewed',
                        'resolved': 'resolved'
                    }
                    status_value = status_map.get(status, 'pending')
                    
                    # Create or update remark
                    remark, created = AttendanceRemark.objects.update_or_create(
                        attendance_record=attendance_record,
                        ep_no=ep_no,
                        date=date,
                        defaults={
                            'reason': reason,
                            'remarks_text': remarks_text,
                            'created_by': created_by,
                            'status': status_value,
                            'admin_response': admin_response
                        }
                    )
                    
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    error_count += 1
            
            # Show results
            if success_count > 0:
                messages.success(request, f'Successfully imported {success_count} remarks.')
            
            if error_count > 0:
                error_msg = f'{error_count} errors occurred. First 10 errors: ' + '; '.join(errors[:10])
                messages.warning(request, error_msg)
            
            logger.info(f"User {request.user.username} uploaded remarks log: {success_count} success, {error_count} errors")
            return redirect('core:remarks_log')
            
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            logger.error(f"Error uploading remarks log: {str(e)}")
            return redirect('core:upload_remarks_log')
    
    # GET request - show upload form
    return render(request, 'upload_remarks_log.html')


# ============================================================================
# Report Views for Different Data Types
# ============================================================================

@login_required
@login_required
@company_access_required
def arc_summary_report(request):
    """ARC Summary Report - Daily Summary Data from AttendanceRecord"""
    from django.core.paginator import Paginator
    from datetime import date
    import calendar
    
    # Get filter parameters
    ep_no = request.GET.get('ep_no', '').strip()
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    dashboard_month = request.GET.get('dashboard_month', '').strip()
    
    # Auto-set date range based on dashboard month or current month
    if not date_from or not date_to:
        if dashboard_month:
            try:
                year, month = map(int, dashboard_month.split('-'))
            except (ValueError, IndexError):
                today = date.today()
                year, month = today.year, today.month
        else:
            today = date.today()
            year, month = today.year, today.month
        
        # Set first day and last day of the month
        first_day = date(year, month, 1)
        last_day = date(year, month, calendar.monthrange(year, month)[1])
        
        if not date_from:
            date_from = first_day.strftime('%Y-%m-%d')
        if not date_to:
            date_to = last_day.strftime('%Y-%m-%d')
    
    # Base queryset - Query AttendanceRecord, exclude records with Unknown contractor
    # Also exclude records with missing essential data
    queryset = AttendanceRecord.objects.select_related('company').exclude(
        cont_code='Unknown'
    ).exclude(
        cont_code=''
    ).exclude(
        company__name='Unknown'
    ).exclude(
        ep_no=''
    ).exclude(
        ep_no__isnull=True
    ).filter(
        company__isnull=False
    )
    
    # Apply company filter based on user role
    if request.user.role == 'admin':
        # Admin sees only their company's data
        if request.user.company:
            queryset = queryset.filter(company=request.user.company)
    elif request.user.role == 'user1':
        # User1 sees only assigned employees
        from core.services.access_control_service import AccessControlService
        access_service = AccessControlService()
        accessible_ep_nos = access_service.get_accessible_ep_numbers(request.user)
        queryset = queryset.filter(ep_no__in=accessible_ep_nos)
    
    # Apply filters
    if ep_no:
        queryset = queryset.filter(ep_no__icontains=ep_no)
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)
    
    # Order by date descending
    queryset = queryset.order_by('-date', 'ep_no')
    
    # Pagination
    paginator = Paginator(queryset, 100)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'ep_no': ep_no,
        'date_from': date_from,
        'date_to': date_to,
        'total_count': paginator.count,
    }
    
    return render(request, 'reports/arc_summary.html', context)


@login_required
@company_access_required
def overtime_report(request):
    """Overtime Report - Records with Overtime requests from AttendanceRecord"""
    from django.core.paginator import Paginator
    from django.db.models import Q
    from datetime import date
    import calendar
    
    # Get filter parameters
    ep_no = request.GET.get('ep_no', '').strip()
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    dashboard_month = request.GET.get('dashboard_month', '').strip()
    
    # Auto-set date range based on dashboard month or current month
    if not date_from or not date_to:
        if dashboard_month:
            try:
                year, month = map(int, dashboard_month.split('-'))
            except (ValueError, IndexError):
                today = date.today()
                year, month = today.year, today.month
        else:
            today = date.today()
            year, month = today.year, today.month
        
        # Set first day and last day of the month
        first_day = date(year, month, 1)
        last_day = date(year, month, calendar.monthrange(year, month)[1])
        
        if not date_from:
            date_from = first_day.strftime('%Y-%m-%d')
        if not date_to:
            date_to = last_day.strftime('%Y-%m-%d')
    
    # Base queryset - Query AttendanceRecord with any overtime request data
    # Show records that have actual_overstay, requested_overtime, or ot_request_status
    queryset = AttendanceRecord.objects.select_related('company').filter(
        Q(actual_overstay__isnull=False, actual_overstay__gt='') |
        Q(requested_overtime__isnull=False, requested_overtime__gt='') |
        Q(ot_request_status__isnull=False, ot_request_status__gt='')
    )
    
    # Apply company filter based on user role
    if request.user.role == 'admin':
        if request.user.company:
            queryset = queryset.filter(company=request.user.company)
    elif request.user.role == 'user1':
        from core.services.access_control_service import AccessControlService
        access_service = AccessControlService()
        accessible_ep_nos = access_service.get_accessible_ep_numbers(request.user)
        queryset = queryset.filter(ep_no__in=accessible_ep_nos)
    
    # Apply filters
    if ep_no:
        queryset = queryset.filter(ep_no__icontains=ep_no)
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)
    
    # Order by date descending
    queryset = queryset.order_by('-date', 'ep_no')
    
    # Pagination
    paginator = Paginator(queryset, 100)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'ep_no': ep_no,
        'date_from': date_from,
        'date_to': date_to,
        'total_count': paginator.count,
    }
    
    return render(request, 'reports/overtime.html', context)


@login_required
@company_access_required
@login_required
@company_access_required
def partial_day_report(request):
    """Partial Day Report - Records with status PD from AttendanceRecord"""
    from django.core.paginator import Paginator
    from datetime import date
    import calendar
    
    # Get filter parameters
    ep_no = request.GET.get('ep_no', '').strip()
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    dashboard_month = request.GET.get('dashboard_month', '').strip()
    
    # Auto-set date range based on dashboard month or current month
    if not date_from or not date_to:
        if dashboard_month:
            try:
                year, month = map(int, dashboard_month.split('-'))
            except (ValueError, IndexError):
                today = date.today()
                year, month = today.year, today.month
        else:
            today = date.today()
            year, month = today.year, today.month
        
        # Set first day and last day of the month
        first_day = date(year, month, 1)
        last_day = date(year, month, calendar.monthrange(year, month)[1])
        
        if not date_from:
            date_from = first_day.strftime('%Y-%m-%d')
        if not date_to:
            date_to = last_day.strftime('%Y-%m-%d')
    
    # Base queryset - Query AttendanceRecord with PD status
    queryset = AttendanceRecord.objects.select_related('company').filter(status='PD')
    
    # Apply company filter based on user role
    if request.user.role == 'admin':
        if request.user.company:
            queryset = queryset.filter(company=request.user.company)
    elif request.user.role == 'user1':
        from core.services.access_control_service import AccessControlService
        access_service = AccessControlService()
        accessible_ep_nos = access_service.get_accessible_ep_numbers(request.user)
        queryset = queryset.filter(ep_no__in=accessible_ep_nos)
    
    # Apply filters
    if ep_no:
        queryset = queryset.filter(ep_no__icontains=ep_no)
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)
    
    # Order by date descending
    queryset = queryset.order_by('-date', 'ep_no')
    
    # Pagination
    paginator = Paginator(queryset, 100)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'ep_no': ep_no,
        'date_from': date_from,
        'date_to': date_to,
        'total_count': paginator.count,
    }
    
    return render(request, 'reports/partial_day.html', context)


@login_required
@company_access_required
def regularization_report(request):
    """Regularization Report - All attendance records (for regularization tracking)"""
    from django.core.paginator import Paginator
    from django.db.models import Q
    from datetime import date
    import calendar
    
    # Get filter parameters
    ep_no = request.GET.get('ep_no', '').strip()
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    dashboard_month = request.GET.get('dashboard_month', '').strip()
    
    # Auto-set date range based on dashboard month or current month
    if not date_from or not date_to:
        if dashboard_month:
            try:
                year, month = map(int, dashboard_month.split('-'))
            except (ValueError, IndexError):
                today = date.today()
                year, month = today.year, today.month
        else:
            today = date.today()
            year, month = today.year, today.month
        
        # Set first day and last day of the month
        first_day = date(year, month, 1)
        last_day = date(year, month, calendar.monthrange(year, month)[1])
        
        if not date_from:
            date_from = first_day.strftime('%Y-%m-%d')
        if not date_to:
            date_to = last_day.strftime('%Y-%m-%d')
    
    # Base queryset - Query all AttendanceRecord
    queryset = AttendanceRecord.objects.select_related('company').all()
    
    # Apply company filter based on user role
    if request.user.role == 'admin':
        if request.user.company:
            queryset = queryset.filter(company=request.user.company)
    elif request.user.role == 'user1':
        from core.services.access_control_service import AccessControlService
        access_service = AccessControlService()
        accessible_ep_nos = access_service.get_accessible_ep_numbers(request.user)
        queryset = queryset.filter(ep_no__in=accessible_ep_nos)
    
    # Apply filters
    if ep_no:
        queryset = queryset.filter(ep_no__icontains=ep_no)
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)
    
    # Exclude records with empty or null status
    queryset = queryset.exclude(Q(ot_request_status='') | Q(ot_request_status__isnull=True))
    
    # Order by date descending
    queryset = queryset.order_by('-date', 'ep_no')
    
    # Pagination
    paginator = Paginator(queryset, 100)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'ep_no': ep_no,
        'date_from': date_from,
        'date_to': date_to,
        'total_count': paginator.count,
    }
    
    return render(request, 'reports/regularization.html', context)


@login_required
@company_access_required
def comprehensive_report(request):
    """Comprehensive Report - Combined view of all attendance data with request statuses"""
    from django.core.paginator import Paginator
    from django.db.models import Q
    from datetime import date
    import calendar
    import logging
    
    logger = logging.getLogger(__name__)
    
    # Get filter parameters
    ep_no = request.GET.get('ep_no', '').strip()
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    dashboard_month = request.GET.get('dashboard_month', '').strip()
    show_incomplete = request.GET.get('show_incomplete', '').strip()
    
    # Auto-set date range based on dashboard month or current month
    if not date_from or not date_to:
        if dashboard_month:
            try:
                year, month = map(int, dashboard_month.split('-'))
            except (ValueError, IndexError):
                today = date.today()
                year, month = today.year, today.month
        else:
            today = date.today()
            year, month = today.year, today.month
        
        # Set first day and last day of the month
        first_day = date(year, month, 1)
        last_day = date(year, month, calendar.monthrange(year, month)[1])
        
        if not date_from:
            date_from = first_day.strftime('%Y-%m-%d')
        if not date_to:
            date_to = last_day.strftime('%Y-%m-%d')
    
    # Base queryset - Get all AttendanceRecord data, excluding incomplete records
    queryset = AttendanceRecord.objects.select_related('company').exclude(
        ep_no=''
    ).exclude(
        ep_no__isnull=True
    ).filter(
        company__isnull=False
    )
    
    # Filter out incomplete records by default (unless explicitly requested)
    if show_incomplete != 'yes':
        # Exclude records that have status 'P' but missing essential time data
        # These are typically records with only status but no actual punch times
        before_count = queryset.count()
        queryset = queryset.exclude(
            Q(status='P') & 
            Q(in_time__isnull=True) & 
            Q(out_time__isnull=True) &
            Q(in_time_2__isnull=True) & 
            Q(out_time_2__isnull=True) &
            Q(in_time_3__isnull=True) & 
            Q(out_time_3__isnull=True)
        )
        after_count = queryset.count()
        logger.info(f'Comprehensive Report: Filtered out {before_count - after_count} incomplete records (show_incomplete={show_incomplete})')
    
    # Apply company filter based on user role
    if request.user.role == 'admin':
        # Admin sees only their company's data
        if request.user.company:
            queryset = queryset.filter(company=request.user.company)
    elif request.user.role == 'user1':
        # User1 sees only assigned employees
        from core.services.access_control_service import AccessControlService
        access_service = AccessControlService()
        accessible_ep_nos = access_service.get_accessible_ep_numbers(request.user)
        queryset = queryset.filter(ep_no__in=accessible_ep_nos)
    
    # Apply filters
    if ep_no:
        queryset = queryset.filter(ep_no__icontains=ep_no)
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)
    
    # Order by date descending
    queryset = queryset.order_by('-date', 'ep_no')
    
    # Debug: Log some sample data and filtering info
    logger.info(f'Comprehensive Report - User: {request.user.username}, Role: {request.user.role}')
    logger.info(f'Filters - EP: "{ep_no}", Date: {date_from} to {date_to}, Show Incomplete: "{show_incomplete}"')
    logger.info(f'Final queryset count: {queryset.count()}')
    
    sample_records = queryset[:3]
    for record in sample_records:
        logger.info(f'Sample record: EP={record.ep_no}, Name={record.get_display_name()}, Date={record.date}, Status={record.status}, In={record.in_time}, Out={record.out_time}')
    
    # Pagination
    paginator = Paginator(queryset, 50)  # Smaller page size due to more columns
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Add additional context for template
    context = {
        'page_obj': page_obj,
        'ep_no': ep_no,
        'date_from': date_from,
        'date_to': date_to,
        'show_incomplete': show_incomplete,
        'total_count': paginator.count,
        'is_filtered': show_incomplete != 'yes',  # Whether incomplete records are filtered out
    }
    
    return render(request, 'reports/comprehensive.html', context)


@login_required
@company_access_required
def comprehensive_report_export_view(request):
    """Export comprehensive report to XLSX with same filters as report view"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.db.models import Q
    from datetime import date
    import calendar
    import logging
    
    logger = logging.getLogger(__name__)
    
    # Get filter parameters (same as comprehensive_report view)
    ep_no = request.GET.get('ep_no', '').strip()
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    dashboard_month = request.GET.get('dashboard_month', '').strip()
    show_incomplete = request.GET.get('show_incomplete', '').strip()
    
    # Auto-set date range based on dashboard month or current month
    if not date_from or not date_to:
        if dashboard_month:
            try:
                year, month = map(int, dashboard_month.split('-'))
            except (ValueError, IndexError):
                today = date.today()
                year, month = today.year, today.month
        else:
            today = date.today()
            year, month = today.year, today.month
        
        # Set first day and last day of the month
        first_day = date(year, month, 1)
        last_day = date(year, month, calendar.monthrange(year, month)[1])
        
        if not date_from:
            date_from = first_day.strftime('%Y-%m-%d')
        if not date_to:
            date_to = last_day.strftime('%Y-%m-%d')
    
    # Base queryset - Get all AttendanceRecord data, excluding incomplete records
    queryset = AttendanceRecord.objects.select_related('company').exclude(
        ep_no=''
    ).exclude(
        ep_no__isnull=True
    ).filter(
        company__isnull=False
    )
    
    # Filter out incomplete records by default (unless explicitly requested)
    if show_incomplete != 'yes':
        # Exclude records that have status 'P' but missing essential time data
        # These are typically records with only status but no actual punch times
        queryset = queryset.exclude(
            Q(status='P') & 
            Q(in_time__isnull=True) & 
            Q(out_time__isnull=True) &
            Q(in_time_2__isnull=True) & 
            Q(out_time_2__isnull=True) &
            Q(in_time_3__isnull=True) & 
            Q(out_time_3__isnull=True)
        )
    
    # Apply company filter based on user role
    if request.user.role == 'admin':
        # Admin sees only their company's data
        if request.user.company:
            queryset = queryset.filter(company=request.user.company)
    elif request.user.role == 'user1':
        # User1 sees only assigned employees
        from core.services.access_control_service import AccessControlService
        access_service = AccessControlService()
        accessible_ep_nos = access_service.get_accessible_ep_numbers(request.user)
        queryset = queryset.filter(ep_no__in=accessible_ep_nos)
    
    # Apply filters
    if ep_no:
        queryset = queryset.filter(ep_no__icontains=ep_no)
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)
    
    # Order by date descending
    queryset = queryset.order_by('-date', 'ep_no')
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Comprehensive Report"
    
    # Define headers
    headers = [
        'EP NO', 'NAME', 'DATE', 'SHIFT', 'IN TIME', 'OUT TIME', 'IN TIME 2', 'OUT TIME 2', 
        'IN TIME 3', 'OUT TIME 3', 'HOURS', 'STATUS', 'REG STATUS', 'PD STATUS', 'OT STATUS',
        'TRADE', 'MANDAYS', 'REG HR', 'OT'
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4A70A9", end_color="4A70A9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    row_num = 2
    for record in queryset:
        # Determine request statuses
        reg_status = "-"
        pd_status = "-"
        ot_status = "-"
        
        if record.in_time_2 or record.out_time_2:
            reg_status = record.ot_request_status or "-"
        
        if record.status == 'PD':
            pd_status = record.ot_request_status or "-"
        
        if record.requested_overtime or record.actual_overstay:
            ot_status = record.ot_request_status or "-"
        
        # Add row data
        row_data = [
            record.ep_no,
            record.get_display_name(),
            record.date.strftime('%d/%m/%Y') if record.date else '',
            record.shift or '-',
            record.in_time.strftime('%H:%M') if record.in_time else '-',
            record.out_time.strftime('%H:%M') if record.out_time else '-',
            record.in_time_2.strftime('%H:%M') if record.in_time_2 else '-',
            record.out_time_2.strftime('%H:%M') if record.out_time_2 else '-',
            record.in_time_3.strftime('%H:%M') if record.in_time_3 else '-',
            record.out_time_3.strftime('%H:%M') if record.out_time_3 else '-',
            record.hours or '-',
            record.status or '-',
            reg_status,
            pd_status,
            ot_status,
            record.trade or '-',
            str(record.mandays) if record.mandays else '0.00',
            record.regular_manday_hr or '-',
            record.ot or '-'
        ]
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=value)
        
        row_num += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    from django.http import HttpResponse
    from datetime import datetime
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="comprehensive_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    
    logger.info(f'User {request.user.username} exported comprehensive report with {queryset.count()} records')
    
    return response
