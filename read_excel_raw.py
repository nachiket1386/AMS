from openpyxl import load_workbook
import sys

try:
    # Read the Excel file without styles
    excel_file = 'CrystalReportViewer1.xlsx'
    
    print("=" * 80)
    print(f"EXCEL FILE ANALYSIS: {excel_file}")
    print("=" * 80)
    
    # Load workbook without data_only and without styles
    wb = load_workbook(excel_file, data_only=True, read_only=True)
    
    print(f"\nTotal Sheets: {len(wb.sheetnames)}")
    print(f"Sheet Names: {wb.sheetnames}")
    print("\n" + "=" * 80)
    
    # Read each sheet
    for sheet_name in wb.sheetnames:
        print(f"\n📊 SHEET: {sheet_name}")
        print("-" * 80)
        
        ws = wb[sheet_name]
        
        # Get dimensions
        print(f"Max Row: {ws.max_row}")
        print(f"Max Column: {ws.max_column}")
        
        # Get headers (first row)
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        print(f"\nColumn Headers ({len(headers)} columns):")
        for i, header in enumerate(headers, 1):
            print(f"  {i}. {header}")
        
        # Read first 10 rows of data
        print(f"\nFirst 10 Rows of Data:")
        print("-" * 80)
        
        row_count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=11, values_only=True), 1):
            if row_idx == 1:
                # Header row
                print("ROW 1 (HEADERS):")
                for i, val in enumerate(row, 1):
                    print(f"  Col {i}: {val}")
            else:
                # Data rows
                print(f"\nROW {row_idx}:")
                for i, val in enumerate(row, 1):
                    if val is not None:
                        print(f"  {headers[i-1] if i-1 < len(headers) else f'Col {i}'}: {val}")
                row_count += 1
        
        # Sample last 5 rows
        print(f"\n\nLast 5 Rows of Data:")
        print("-" * 80)
        
        start_row = max(2, ws.max_row - 4)
        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start_row):
            print(f"\nROW {row_idx}:")
            for i, val in enumerate(row, 1):
                if val is not None:
                    print(f"  {headers[i-1] if i-1 < len(headers) else f'Col {i}'}: {val}")
        
        # Count non-empty rows
        non_empty_rows = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                non_empty_rows += 1
        
        print(f"\n\nTotal Data Rows (excluding header): {non_empty_rows}")
        
        print("\n" + "=" * 80)
    
    wb.close()

except Exception as e:
    print(f"Error reading Excel file: {e}")
    import traceback
    traceback.print_exc()
